"""
Task AI Summary — gathers task context and calls Gemini to produce a
structured summary and answer user questions about a specific task.
"""
import base64
import json
import logging
import os

import requests
from django.conf import settings

from .models import TaskComment, TaskFieldHistory, TaskRelation

logger = logging.getLogger(__name__)

_GEMINI_MODEL = "gemini-2.5-flash-lite"
_CACHE_TTL = 300  # 5 minutes (optimal breakeven at 1 follow-up question)
_REDIS_KEY_PREFIX = "task_ai_cache"


def _get_api_key() -> str:
    api_key = (
        getattr(settings, "GEMINI_API_KEY", "")
        or os.environ.get("GEMINI_API_KEY", "")
    )
    if not api_key:
        raise RuntimeError("GEMINI_API_KEY is not configured")
    return api_key


def invalidate_task_ai_cache(task_id: int) -> None:
    """Invalidate the Gemini Q&A cache after any task mutation."""
    from django.core.cache import cache
    cache.delete(f"{_REDIS_KEY_PREFIX}:{task_id}")
    logger.info("Invalidated AI cache for task %s", task_id)

# Max characters to extract per text-based attachment (plain text, PDF, Word)
_MAX_ATTACHMENT_CHARS = 50_000
# Max rows per sheet for Excel — prevents enormous data exports from timing out
_MAX_ROWS_PER_SHEET = 10_000
# Max file size pre-check for documents
_MAX_DOCUMENT_BYTES = 20 * 1024 * 1024   # 20 MB
_MAX_TEXT_BYTES = 5 * 1024 * 1024        # 5 MB for plain text/CSV/JSON
# Max raw bytes to send per image attachment (10 MB)
_MAX_IMAGE_BYTES = 10 * 1024 * 1024


def _call_gemini(
    system_prompt: str,
    user_prompt: str,
    timeout: int = 55,
    image_parts: list | None = None,
) -> str:
    """
    Direct non-streaming Gemini call via generateContent.
    Optionally accepts image_parts for vision tasks.
    """
    api_key = _get_api_key()
    url = (
        f"https://aiplatform.googleapis.com/v1/publishers/google/models/"
        f"{_GEMINI_MODEL}:generateContent?key={api_key}"
    )
    user_parts: list = [{"text": user_prompt}]
    if image_parts:
        user_parts.extend(image_parts)
    body = {
        "contents": [{"role": "user", "parts": user_parts}],
        "systemInstruction": {"parts": [{"text": system_prompt}]},
        "generationConfig": {"temperature": 0.2},
    }
    resp = requests.post(url, json=body, timeout=timeout)
    resp.raise_for_status()
    data = resp.json()
    parts = (
        data.get("candidates", [{}])[0]
        .get("content", {})
        .get("parts", [])
    )
    return "".join(p.get("text", "") for p in parts).strip()


# Both cache creation and cached generation use the AI Studio endpoint.
# Vertex AI keys (aiplatform.googleapis.com) do NOT work here — they require
# project/location OAuth2. If GEMINI_API_KEY is a Vertex AI key, caching will
# fail gracefully and fall back to non-cached calls every time.
# To enable caching, set GEMINI_AI_STUDIO_KEY to an AI Studio API key.
_AI_STUDIO_BASE = "https://generativelanguage.googleapis.com/v1beta"


def _get_cache_api_key() -> "str | None":
    """Return an AI Studio API key for caching, or None if not configured."""
    key = (
        getattr(settings, "GEMINI_AI_STUDIO_KEY", "")
        or os.environ.get("GEMINI_AI_STUDIO_KEY", "")
        # Fall back to the main key — works if the user has an AI Studio key set as GEMINI_API_KEY
        or getattr(settings, "GEMINI_API_KEY", "")
        or os.environ.get("GEMINI_API_KEY", "")
    )
    return key or None


def _create_gemini_cache(
    system_prompt: str,
    context_text: str,
    image_parts: list,
    ttl: int = _CACHE_TTL,
) -> "str | None":
    """
    Create a Gemini server-side cache via the AI Studio endpoint.
    Returns the cache name or None if unsupported/failed.
    Minimum 1,024 tokens required — silently returns None for small contexts.
    """
    api_key = _get_cache_api_key()
    if not api_key:
        return None

    url = f"{_AI_STUDIO_BASE}/cachedContents?key={api_key}"
    user_parts: list = [{"text": context_text}]
    if image_parts:
        user_parts.extend(image_parts)
    body = {
        "model": f"models/{_GEMINI_MODEL}",
        "systemInstruction": {"parts": [{"text": system_prompt}]},
        "contents": [{"role": "user", "parts": user_parts}],
        "ttl": f"{ttl}s",
    }
    try:
        resp = requests.post(url, json=body, timeout=30)
        if resp.status_code in (401, 403):
            logger.warning(
                "Gemini caching requires an AI Studio API key (GEMINI_AI_STUDIO_KEY). "
                "Your current key appears to be a Vertex AI key which does not support "
                "the caching API. Falling back to non-cached calls."
            )
            return None
        resp.raise_for_status()
        name = resp.json().get("name")
        logger.info("Created Gemini cache: %s (TTL=%ds)", name, ttl)
        return name
    except Exception as exc:
        logger.info("Gemini cache creation skipped: %s", exc)
        return None


def _call_gemini_cached(cache_name: str, question: str, timeout: int = 55) -> str:
    """Answer a question using a Gemini server-side cache via AI Studio endpoint."""
    api_key = _get_cache_api_key()
    url = f"{_AI_STUDIO_BASE}/models/{_GEMINI_MODEL}:generateContent?key={api_key}"
    body = {
        "cachedContent": cache_name,
        "contents": [{"role": "user", "parts": [{"text": question}]}],
        "generationConfig": {"temperature": 0.2},
    }
    resp = requests.post(url, json=body, timeout=timeout)
    resp.raise_for_status()
    data = resp.json()
    parts = (
        data.get("candidates", [{}])[0]
        .get("content", {})
        .get("parts", [])
    )
    return "".join(p.get("text", "") for p in parts).strip()


_SUMMARY_SYSTEM_PROMPT = """
You are an AI assistant that analyzes task management records.
Given a task's full context (details, comments, field changes, related tasks, and
optionally attachment contents or images), produce a structured JSON summary.
Be concise — 1–2 sentences per item.

Return ONLY valid JSON with this exact schema:
{
  "key_decisions": [],
  "blockers": [],
  "escalations": [],
  "action_items": [],
  "status_changes": [],
  "unresolved_questions": []
}

- key_decisions: Decisions made or agreed upon in comments, history, or attachments
- blockers: Current impediments or blocking issues
- escalations: Issues that have been or need to be escalated
- action_items: Specific tasks or next steps assigned to people
- status_changes: Notable status, priority, or ownership changes with context
- unresolved_questions: Open questions that have not been answered yet

Return empty arrays for sections with nothing relevant.
""".strip()

_QA_SYSTEM_PROMPT = """
You are an AI assistant with full context of a task, including any uploaded attachments.
Answer the user's question concisely and accurately based only on the provided context.
If the answer is not in the context, say so clearly.
""".strip()


# ── Attachment extraction ──────────────────────────────────────────────────


def _extract_attachment_content(att) -> "tuple[str | dict | None, bool]":
    """
    Extract readable content from a clean attachment.

    Returns (content, truncated):
      content:   str (text), dict (Gemini image inline_data), or None (unsupported/failed)
      truncated: True if text content was cut short due to size limits
    """
    ct = (att.content_type or "").lower()

    try:
        file_path = att.file.path
        file_size = att.file_size or 0

        # ── CSV (treated as tabular data — no char cap, like Excel) ─────────
        if ct in ("text/csv", "application/csv") or att.original_filename.lower().endswith(".csv"):
            if file_size > _MAX_DOCUMENT_BYTES:
                logger.info("CSV attachment %s too large (%d bytes), skipping", att.original_filename, file_size)
                return None, False
            with open(file_path, "r", errors="replace") as f:
                text = f.read()
            return text, False

        # ── Plain text / JSON / XML / HTML ────────────────────────────────
        elif ct.startswith("text/") or ct in (
            "application/json",
            "application/xml",
        ):
            if file_size > _MAX_TEXT_BYTES:
                logger.info("Text attachment %s too large (%d bytes), skipping", att.original_filename, file_size)
                return None, False
            with open(file_path, "r", errors="replace") as f:
                text = f.read(_MAX_ATTACHMENT_CHARS + 1)
            truncated = len(text) > _MAX_ATTACHMENT_CHARS
            return text[:_MAX_ATTACHMENT_CHARS], truncated

        # ── PDF ───────────────────────────────────────────────────────────
        elif ct == "application/pdf":
            if file_size > _MAX_DOCUMENT_BYTES:
                logger.info("PDF attachment %s too large (%d bytes), skipping", att.original_filename, file_size)
                return None, False
            try:
                import fitz  # pymupdf
                doc = fitz.open(file_path)
                text = ""
                truncated = False
                for page in doc:
                    text += page.get_text()
                    if len(text) >= _MAX_ATTACHMENT_CHARS:
                        truncated = True
                        break
                doc.close()
                return text[:_MAX_ATTACHMENT_CHARS], truncated
            except ImportError:
                logger.warning("pymupdf not installed; skipping PDF attachment %s", att.original_filename)
                return None, False

        # ── Word (.docx) ──────────────────────────────────────────────────
        elif ct in (
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            "application/msword",
        ):
            if file_size > _MAX_DOCUMENT_BYTES:
                logger.info("Word attachment %s too large (%d bytes), skipping", att.original_filename, file_size)
                return None, False
            try:
                from docx import Document
                doc = Document(file_path)
                text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
                truncated = len(text) > _MAX_ATTACHMENT_CHARS
                return text[:_MAX_ATTACHMENT_CHARS], truncated
            except ImportError:
                logger.warning("python-docx not installed; skipping Word attachment %s", att.original_filename)
                return None, False

        # ── Excel (.xlsx / .xls) ──────────────────────────────────────────
        elif ct in (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
        ):
            if file_size > _MAX_DOCUMENT_BYTES:
                logger.info("Excel attachment %s too large (%d bytes), skipping", att.original_filename, file_size)
                return None, False
            try:
                import openpyxl
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                sections = []
                truncated = False
                for sheet in wb.worksheets:
                    sheet_lines = [f"[Sheet: {sheet.title}]"]
                    for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                        if row_idx >= _MAX_ROWS_PER_SHEET:
                            sheet_lines.append(f"… ({sheet.max_row - _MAX_ROWS_PER_SHEET} more rows truncated)")
                            truncated = True
                            break
                        line = "\t".join("" if c is None else str(c) for c in row)
                        sheet_lines.append(line)
                    sections.append("\n".join(sheet_lines))
                wb.close()
                return "\n\n".join(sections), truncated
            except ImportError:
                logger.warning("openpyxl not installed; skipping Excel attachment %s", att.original_filename)
                return None, False

        # ── Images ────────────────────────────────────────────────────────
        elif ct.startswith("image/"):
            if file_size > _MAX_IMAGE_BYTES:
                logger.info("Image attachment %s too large (%d bytes), skipping", att.original_filename, file_size)
                return None, False
            with open(file_path, "rb") as f:
                raw = f.read()
            return {
                "inline_data": {
                    "mime_type": ct,
                    "data": base64.b64encode(raw).decode(),
                }
            }, False

    except Exception as exc:
        logger.warning(
            "Could not extract attachment %s: %s", att.original_filename, exc
        )

    return None, False


# ── Context builder ────────────────────────────────────────────────────────


def _gather_context(task) -> "tuple[str, list, list]":
    """
    Build a text context string and a list of Gemini image parts for the task.

    Returns:
      (context_text, image_parts, attachment_meta)
    """
    from .models import TaskAttachment  # avoid circular at module level

    lines = []
    image_parts: list = []

    lines.append(f"=== TASK #{task.id} ===")
    lines.append(f"Summary: {task.summary}")
    lines.append(f"Type: {task.type}")
    lines.append(f"Status: {task.status}")
    lines.append(f"Priority: {task.priority or 'Not set'}")
    lines.append(f"Due date: {task.due_date or 'Not set'}")
    lines.append(f"Creator: {task.created_by.username if task.created_by else 'Unknown'}")
    lines.append(f"Owner: {task.owner.username if task.owner else 'Unassigned'}")
    lines.append(
        f"Current approver: {task.current_approver.username if task.current_approver else 'None'}"
    )
    if task.description:
        lines.append(f"Description: {task.description}")

    comments = (
        TaskComment.objects.filter(task=task)
        .select_related("user")
        .order_by("created_at")[:50]
    )
    if comments:
        lines.append("\n=== COMMENTS ===")
        for c in comments:
            lines.append(
                f"[{c.created_at.strftime('%Y-%m-%d %H:%M')}] {c.user.username}: {c.body}"
            )

    history = (
        TaskFieldHistory.objects.filter(task=task)
        .select_related("changed_by")
        .order_by("-changed_at")[:30]
    )
    if history:
        lines.append("\n=== FIELD CHANGES (recent first) ===")
        for h in history:
            by = h.changed_by.username if h.changed_by else "system"
            lines.append(
                f"[{h.changed_at.strftime('%Y-%m-%d %H:%M')}] {by} changed {h.field_name}: "
                f"{h.old_value!r} → {h.new_value!r}"
            )

    relations = TaskRelation.objects.filter(
        source_task=task
    ).select_related("target_task").order_by("relationship_type")
    if relations:
        lines.append("\n=== RELATIONS ===")
        for r in relations:
            lines.append(
                f"{r.relationship_type} → Task #{r.target_task_id}: {r.target_task.summary}"
            )

    # ── Attachments ────────────────────────────────────────────────────────
    attachments = (
        TaskAttachment.objects.filter(task=task, scan_status="clean")
        .select_related("uploaded_by")
        .order_by("created_at")
    )
    attachment_meta = []
    if attachments:
        lines.append("\n=== ATTACHMENTS ===")
        for att in attachments:
            uploader = att.uploaded_by.username if att.uploaded_by else "unknown"
            lines.append(
                f"- {att.original_filename} "
                f"({att.content_type}, {att.file_size} bytes, "
                f"uploaded by {uploader} on {att.created_at.strftime('%Y-%m-%d')})"
            )
            content, truncated = _extract_attachment_content(att)
            if content is None:
                lines.append("  [Content not extractable]")
                attachment_meta.append({
                    "name": att.original_filename,
                    "type": att.content_type or "unknown",
                    "truncated": False,
                    "readable": False,
                })
            elif isinstance(content, dict):
                image_parts.append(content)
                lines.append("  [Image — sent to AI vision]")
                attachment_meta.append({
                    "name": att.original_filename,
                    "type": att.content_type or "unknown",
                    "truncated": False,
                    "readable": True,
                })
            else:
                lines.append(f"  Content:\n{content}")
                attachment_meta.append({
                    "name": att.original_filename,
                    "type": att.content_type or "unknown",
                    "truncated": truncated,
                    "readable": True,
                })

    return "\n".join(lines), image_parts, attachment_meta


# ── Public API ─────────────────────────────────────────────────────────────


def generate_ai_summary(task) -> dict:
    context, image_parts, attachment_meta = _gather_context(task)
    user_prompt = f"Summarize the following task:\n\n{context}"
    try:
        raw = _call_gemini(_SUMMARY_SYSTEM_PROMPT, user_prompt, image_parts=image_parts or None)
        # strip markdown fences if present
        clean = raw.strip()
        if clean.startswith("```"):
            clean = "\n".join(clean.splitlines()[1:])
            clean = clean.rstrip("`").strip()
        try:
            result = json.loads(clean)
        except json.JSONDecodeError:
            start = clean.find("{")
            end = clean.rfind("}") + 1
            result = json.loads(clean[start:end])
        for key in (
            "key_decisions",
            "blockers",
            "escalations",
            "action_items",
            "status_changes",
            "unresolved_questions",
        ):
            if key not in result or not isinstance(result[key], list):
                result[key] = []
            else:
                # Flatten any objects Gemini returns instead of plain strings
                result[key] = [
                    item
                    if isinstance(item, str)
                    else " — ".join(f"{v}" for v in item.values())
                    if isinstance(item, dict)
                    else str(item)
                    for item in result[key]
                ]
        result["_attachments"] = attachment_meta
        return result
    except Exception as exc:
        logger.error("AI summary failed for task %s: %s", task.id, exc)
        raise


def answer_task_question(task, question: str) -> str:
    from django.core.cache import cache as django_cache

    redis_key = f"{_REDIS_KEY_PREFIX}:{task.id}"
    gemini_cache_name = django_cache.get(redis_key)

    # ── Try cached path ────────────────────────────────────────────────────
    if gemini_cache_name:
        try:
            logger.info("Using Gemini cache %s for task %s", gemini_cache_name, task.id)
            return _call_gemini_cached(gemini_cache_name, f"=== QUESTION ===\n{question}")
        except Exception as exc:
            logger.warning(
                "Gemini cached call failed for task %s, rebuilding context: %s", task.id, exc
            )
            django_cache.delete(redis_key)

    # ── Build full context ─────────────────────────────────────────────────
    context_text, image_parts, _ = _gather_context(task)

    # Try to create a cache for all future questions in this session
    gemini_cache_name = _create_gemini_cache(_QA_SYSTEM_PROMPT, context_text, image_parts)
    if gemini_cache_name:
        # Redis TTL is slightly longer than Gemini's to avoid the edge case
        # where Redis expires a second before Gemini and triggers a needless rebuild
        django_cache.set(redis_key, gemini_cache_name, _CACHE_TTL + 60)

    # Answer this question with the full context (first call is always full price)
    user_prompt = f"{context_text}\n\n=== QUESTION ===\n{question}"
    try:
        return _call_gemini(_QA_SYSTEM_PROMPT, user_prompt, image_parts=image_parts or None)
    except Exception as exc:
        logger.error("AI Q&A failed for task %s: %s", task.id, exc)
        raise
